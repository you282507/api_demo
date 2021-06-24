from bns.api.dev import api_dev
from bns.api.web_rw.api_web_rw import *
import openpyxl


def get_dev_alarm_type(vehicle_no, data_day):
	alarm_type_dict = dict()
	current = 1
	while True:
		all_data = get_alarm_info(vehicle_no, data_day, current)
		alarm_data_list = all_data['response_data']["data"]["data"]
		if len(alarm_data_list) == 0:
			break
		for data in alarm_data_list:
			alarm_type = data["alarmTypeValue"]
			if alarm_type in alarm_type_dict.keys():
				alarm_type_dict[alarm_type] = alarm_type_dict[alarm_type] + 1
			else:
				alarm_type_dict[alarm_type] = 1
		current += 1
	return alarm_type_dict


class DoExcel:
	def __init__(self, file_path):
		self.file_path = file_path
		self.obj = openpyxl.load_workbook(file_path)
		self.sheet = self.obj.active
		self.title_dict = self.write_title()

	def read_excel(self, row, column):
		return self.sheet.cell(row, column).value

	def write_excel(self, row, column, value):
		"""
		Write data to the specified location
		@param row:line
		@param column:column
		@param value:write str
		"""
		self.sheet.cell(row, column).value = value
		self.obj.save(self.file_path)
		self.obj.close()

	def get_excel_row(self):
		return self.sheet.max_row

	def write_title(self, start_column=11):
		title_list = [
			'分神驾驶', '前车碰撞', '驾驶员异常', '超时疲劳驾驶预警',
			'行人碰撞', '驾驶员变更', '疲劳驾驶', '抽烟', '超速预警',
			'驾驶辅助功能失效报警', '红外阻断报警','驾驶员行为监测功能失效报警',
			'接打电话', '车辆非法位移', '疲劳驾驶报警', '车道偏离', '超时停车报警', '超速报警', '车距过近']
		title_dict = dict()
		for title in title_list:
			self.sheet.cell(1, start_column).value = title
			title_dict[title] = start_column
			start_column += 1
		self.obj.save(self.file_path)
		self.obj.close()
		return title_dict

	def write_alarm(self, row, alarm_info):
		for alarm_name in alarm_info.keys():
			column = self.title_dict[alarm_name]
			self.write_excel(row, column, alarm_info[alarm_name])


def get_alarm_to_excel(file_path):
	do_excel_obj = DoExcel(file_path)
	excel_row = do_excel_obj.get_excel_row()
	for o in range(excel_row - 1):
		r = o + 2
		data_time = do_excel_obj.read_excel(r, 1)
		dev_no = do_excel_obj.read_excel(r, 2)
		alarm_info = get_dev_alarm_type(dev_no, data_time)
		do_excel_obj.write_alarm(r, alarm_info)


if __name__ == '__main__':
	a = "result/河南中交-2021-01-133.xlsx"
	get_alarm_to_excel(a)